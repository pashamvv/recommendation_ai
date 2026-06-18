from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
import logging
from pathlib import Path
import re
from typing import Any

from sqlalchemy import delete, insert
from sqlalchemy.orm import Session

from models import Actor, Director, Genre, Keyword, Movie, movie_actors
from tmdb_service import TMDBServiceError, tmdb_service


class ExcelServiceError(RuntimeError):
    pass


logger = logging.getLogger(__name__)


class ExcelService:
    export_dir = Path("exports")

    def _load_openpyxl(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as exc:
            raise ExcelServiceError(
                "openpyxl is not installed. Run pip install -r requirements.txt.",
            ) from exc
        return Workbook, load_workbook

    def _resolve_export_path(self, file_name: str | None = None) -> Path:
        self.export_dir.mkdir(parents=True, exist_ok=True)
        if not file_name:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            file_name = f"tmdb_export_{timestamp}.xlsx"
        if not file_name.endswith(".xlsx"):
            file_name = f"{file_name}.xlsx"
        return (self.export_dir / file_name).resolve()

    def _resolve_import_path(self, file_path: str) -> Path:
        path = Path(file_path).expanduser()
        if not path.is_absolute():
            path = (Path.cwd() / path).resolve()
        if not path.exists():
            raise ExcelServiceError(f"Excel file not found: {path}")
        return path

    def _flat_movie_headers(self) -> list[str]:
        return [
            "tmdb_id",
            "title",
            "description",
            "release_year",
            "poster_path",
            "tmdb_popularity",
            "tmdb_vote",
            "genres",
            "tags",
            "countries",
            "languages",
            "cast",
            "crew",
        ]

    def _pipe_join(self, values: list[str]) -> str:
        cleaned = [value.strip() for value in values if value and value.strip()]
        return "|".join(dict.fromkeys(cleaned))

    def _normalize_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9_]+", "", value.strip().lower().replace(" ", "_"))

    def _normalize_image_value(self, value: str | None) -> str | None:
        if not value:
            return None
        if value.startswith("http://") or value.startswith("https://"):
            return value
        if value.startswith("/"):
            return tmdb_service.build_image_url(value)
        return value

    def _release_year_to_date(self, value: Any):
        if value in (None, ""):
            return None
        try:
            year = int(value)
        except (TypeError, ValueError):
            return self._as_date(value)
        return date(year, 1, 1)

    def _parse_list_cell(self, value: Any) -> list[str]:
        if value in (None, ""):
            return []
        return [item.strip() for item in str(value).split("|") if item and item.strip()]

    def _parse_crew_cell(self, value: Any) -> list[tuple[str, str | None]]:
        items = []
        for raw_item in self._parse_list_cell(value):
            match = re.match(r"^(?P<name>.+?)\s*\((?P<role>.+?)\)\s*$", raw_item)
            if match:
                items.append((match.group("name").strip(), match.group("role").strip()))
            else:
                items.append((raw_item, None))
        return items

    def _flat_movie_export_row(self, details: dict, credits: dict, keywords_payload: list[dict]) -> list[Any]:
        cast_names = self._pipe_join(
            [
                actor.get("name") or actor.get("original_name") or "Unknown"
                for actor in credits.get("cast", [])[:8]
            ],
        )
        crew_items = self._pipe_join(
            [
                f"{member.get('name') or member.get('original_name') or 'Unknown'} ({member.get('job')})"
                for member in credits.get("crew", [])
                if member.get("job") in {"Director", "Writer", "Screenplay", "Producer", "Executive Producer"}
            ],
        )
        genres = self._pipe_join([genre.get("name") for genre in details.get("genres", [])])
        tags = self._pipe_join([keyword.get("name") for keyword in keywords_payload])
        countries = self._pipe_join(
            [country.get("name") for country in details.get("production_countries", [])],
        )
        languages = self._pipe_join(
            [
                language.get("english_name") or language.get("name") or language.get("iso_639_1")
                for language in details.get("spoken_languages", [])
            ],
        )
        release_date = details.get("release_date") or ""
        release_year = None
        if release_date:
            parsed_date = tmdb_service.parse_date(release_date)
            release_year = parsed_date.year if parsed_date else None

        return [
            details.get("id"),
            details.get("title") or details.get("name") or "Untitled",
            details.get("overview"),
            release_year,
            details.get("poster_path"),
            details.get("popularity") or 0,
            details.get("vote_average") or 0,
            genres,
            tags,
            countries,
            languages or details.get("original_language"),
            cast_names,
            crew_items,
        ]

    def _flat_movie_export_row_from_source_item(
        self,
        item: dict,
        genre_map: dict[int, str],
    ) -> list[Any]:
        release_date = item.get("release_date") or ""
        release_year = None
        if release_date:
            parsed_date = tmdb_service.parse_date(release_date)
            release_year = parsed_date.year if parsed_date else None

        genres = self._pipe_join(
            [genre_map.get(int(genre_id), str(genre_id)) for genre_id in item.get("genre_ids", [])],
        )
        language_value = item.get("original_language") or ""

        return [
            item.get("id"),
            item.get("title") or item.get("name") or "Untitled",
            item.get("overview"),
            release_year,
            item.get("poster_path"),
            item.get("popularity") or 0,
            item.get("vote_average") or 0,
            genres,
            "",
            "",
            language_value,
            "",
            "",
        ]

    def export_tmdb_to_excel(
        self,
        source: str,
        pages: int,
        language: str,
        file_name: str | None = None,
    ) -> dict[str, Any]:
        Workbook, _ = self._load_openpyxl()
        workbook = Workbook()

        movies_sheet = workbook.active
        movies_sheet.title = "movies"
        movies_sheet.append(self._flat_movie_headers())
        movies_sheet.freeze_panes = "A2"
        movies_sheet.auto_filter.ref = "A1:M1"
        movies_sheet.column_dimensions["B"].width = 24
        movies_sheet.column_dimensions["C"].width = 48
        movies_sheet.column_dimensions["H"].width = 24
        movies_sheet.column_dimensions["I"].width = 28
        movies_sheet.column_dimensions["J"].width = 24
        movies_sheet.column_dimensions["K"].width = 20
        movies_sheet.column_dimensions["L"].width = 32
        movies_sheet.column_dimensions["M"].width = 64

        exported_movie_ids: set[int] = set()
        skipped_movies: list[tuple[int, str]] = []
        try:
            genre_map = tmdb_service.fetch_genre_map(language=language)
        except TMDBServiceError:
            genre_map = {}
        for page in range(1, pages + 1):
            payload = tmdb_service.fetch_source_page(source=source, page=page, language=language)
            for item in payload.get("results", []):
                tmdb_id = item["id"]
                if tmdb_id in exported_movie_ids:
                    continue

                try:
                    bundle = tmdb_service.get_movie_export_bundle(tmdb_id=tmdb_id, language=language)
                    details = bundle["details"]
                    credits = bundle["credits"]
                    keywords = bundle["keywords"].get("keywords") or bundle["keywords"].get("results", [])
                    movies_sheet.append(self._flat_movie_export_row(details, credits, keywords))
                    exported_movie_ids.add(tmdb_id)
                except TMDBServiceError as exc:
                    logger.warning("Falling back to source-page data for TMDB movie %s during export: %s", tmdb_id, exc)
                    movies_sheet.append(self._flat_movie_export_row_from_source_item(item, genre_map))
                    exported_movie_ids.add(tmdb_id)
                    skipped_movies.append((tmdb_id, str(exc)))

        if not exported_movie_ids:
            if skipped_movies:
                last_tmdb_id, last_error = skipped_movies[-1]
                raise ExcelServiceError(
                    f"Could not export movies from TMDB. Last failed movie {last_tmdb_id}: {last_error}",
                )
            raise ExcelServiceError("Could not export movies from TMDB.")

        exported_at = datetime.utcnow().isoformat()
        workbook.active = 0
        file_path = self._resolve_export_path(file_name)
        workbook.save(file_path)
        return {
            "source": source,
            "pages": pages,
            "language": language,
            "file_path": str(file_path),
            "exported_movies": len(exported_movie_ids),
            "exported_at": exported_at,
        }

    def _sheet_rows(self, workbook, sheet_name: str) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            return []

        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [self._normalize_header(str(header)) if header is not None else "" for header in rows[0]]
        return [
            {
                headers[index]: row[index]
                for index in range(len(headers))
                if headers[index]
            }
            for row in rows[1:]
            if any(value is not None and value != "" for value in row)
        ]

    def _as_int(self, value: Any) -> int | None:
        if value in (None, ""):
            return None
        return int(value)

    def _as_float(self, value: Any) -> float:
        if value in (None, ""):
            return 0.0
        return float(value)

    def _as_date(self, value: Any):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return tmdb_service.parse_date(str(value))

    def _existing_movie_tmdb_ids(self, db: Session, tmdb_ids: set[int]) -> set[int]:
        if not tmdb_ids:
            return set()
        return {
            tmdb_id
            for (tmdb_id,) in db.query(Movie.tmdb_id).filter(Movie.tmdb_id.in_(tmdb_ids)).all()
        }

    def _insert_movie_from_row(self, db: Session, row: dict[str, Any], id_key: str) -> Movie:
        tmdb_id = self._as_int(row.get(id_key))
        if tmdb_id is None:
            raise ExcelServiceError("Movie row is missing tmdb_id.")

        movie = Movie(tmdb_id=tmdb_id)
        db.add(movie)
        movie.title = row.get("title") or "Untitled"
        movie.original_title = row.get("original_title")
        movie.overview = row.get("overview") or row.get("description")
        movie.release_date = self._as_date(row.get("release_date"))
        movie.rating = self._as_float(row.get("rating"))
        movie.popularity = self._as_float(row.get("popularity"))
        movie.poster_url = self._normalize_image_value(row.get("poster_url"))
        movie.backdrop_url = self._normalize_image_value(row.get("backdrop_url"))
        movie.runtime = self._as_int(row.get("runtime"))
        movie.status = row.get("status")
        movie.language = row.get("language")
        db.flush()
        return movie

    def _insert_movie_from_flat_row(self, db: Session, row: dict[str, Any]) -> Movie:
        tmdb_id = self._as_int(row.get("tmdb_id"))
        if tmdb_id is None:
            raise ExcelServiceError("Movie row is missing tmdb_id.")

        movie = Movie(tmdb_id=tmdb_id)
        db.add(movie)
        movie.title = row.get("title") or "Untitled"
        movie.overview = row.get("description")
        movie.release_date = self._release_year_to_date(row.get("release_year"))
        movie.rating = self._as_float(row.get("tmdb_vote"))
        movie.popularity = self._as_float(row.get("tmdb_popularity"))
        movie.poster_url = self._normalize_image_value(row.get("poster_path"))
        movie.language = (self._parse_list_cell(row.get("languages")) or [None])[0]
        db.flush()
        return movie

    def _get_or_create_genre(self, db: Session, name: str) -> Genre:
        genre = db.query(Genre).filter(Genre.name == name).first()
        if genre is None:
            genre = Genre(name=name)
            db.add(genre)
            db.flush()
        return genre

    def _get_or_create_keyword(self, db: Session, name: str) -> Keyword:
        keyword = db.query(Keyword).filter(Keyword.name == name).first()
        if keyword is None:
            keyword = Keyword(name=name)
            db.add(keyword)
            db.flush()
        return keyword

    def _get_or_create_actor(self, db: Session, name: str) -> Actor:
        actor = db.query(Actor).filter(Actor.name == name).first()
        if actor is None:
            actor = Actor(name=name)
            db.add(actor)
            db.flush()
        return actor

    def _get_or_create_director(self, db: Session, name: str) -> Director:
        director = db.query(Director).filter(Director.name == name).first()
        if director is None:
            director = Director(name=name)
            db.add(director)
            db.flush()
        return director

    def _import_flat_movies_sheet(self, db: Session, movie_rows: list[dict[str, Any]]) -> dict[str, Any]:
        source_movies: dict[int, Movie] = {}
        imported_movie_ids: set[int] = set()
        movie_tmdb_ids = {
            tmdb_id
            for tmdb_id in (self._as_int(row.get("tmdb_id")) for row in movie_rows)
            if tmdb_id is not None
        }
        existing_movie_tmdb_ids = self._existing_movie_tmdb_ids(db, movie_tmdb_ids)
        known_movie_tmdb_ids = set(existing_movie_tmdb_ids)
        skipped_existing_movies = 0

        for row in movie_rows:
            tmdb_id = self._as_int(row.get("tmdb_id"))
            if tmdb_id is None:
                raise ExcelServiceError("Movie row is missing tmdb_id.")
            if tmdb_id in known_movie_tmdb_ids:
                skipped_existing_movies += 1
                continue

            movie = self._insert_movie_from_flat_row(db, row)
            source_movies[movie.tmdb_id] = movie
            imported_movie_ids.add(movie.id)
            known_movie_tmdb_ids.add(tmdb_id)

        for movie_tmdb_id, movie in source_movies.items():
            row = next(item for item in movie_rows if self._as_int(item.get("tmdb_id")) == movie_tmdb_id)

            genre_names = self._parse_list_cell(row.get("genres"))
            movie.genres = [self._get_or_create_genre(db, name) for name in genre_names]

            keyword_names = self._parse_list_cell(row.get("tags"))
            movie.keywords = [self._get_or_create_keyword(db, name) for name in keyword_names]

            cast_names = self._parse_list_cell(row.get("cast"))
            db.execute(delete(movie_actors).where(movie_actors.c.movie_id == movie.id))
            actor_rows_to_insert = []
            for cast_order, actor_name in enumerate(cast_names):
                actor = self._get_or_create_actor(db, actor_name)
                actor_rows_to_insert.append(
                    {
                        "movie_id": movie.id,
                        "actor_id": actor.id,
                        "character_name": None,
                        "cast_order": cast_order,
                    },
                )
            if actor_rows_to_insert:
                db.execute(insert(movie_actors), actor_rows_to_insert)

            directors = []
            for crew_name, crew_role in self._parse_crew_cell(row.get("crew")):
                if crew_role and "director" not in crew_role.lower():
                    continue
                directors.append(self._get_or_create_director(db, crew_name))
            movie.directors = list(dict.fromkeys(directors))

        db.commit()
        return {
            "processed_movies": len(movie_rows),
            "imported_movies": len(source_movies),
            "upserted_movies": len(imported_movie_ids),
            "skipped_existing_movies": skipped_existing_movies,
            "imported_movie_ids": list(imported_movie_ids),
            "imported_at": datetime.utcnow().isoformat(),
        }

    def _import_tmdb_from_workbook(
        self,
        db: Session,
        workbook,
        source_name: str,
    ) -> dict[str, Any]:
        movie_rows = self._sheet_rows(workbook, "movies")
        if not movie_rows:
            raise ExcelServiceError("The Excel file does not contain any movie rows.")

        movie_headers = set(movie_rows[0].keys())
        flat_headers = set(self._flat_movie_headers())
        normalized_flat_headers = {self._normalize_header(header) for header in flat_headers}
        if normalized_flat_headers.issubset(movie_headers):
            result = self._import_flat_movies_sheet(db, movie_rows)
            result["file_path"] = source_name
            return result

        source_movies: dict[int, Movie] = {}
        imported_movie_ids: set[int] = set()
        movie_tmdb_ids = {
            tmdb_id
            for tmdb_id in (self._as_int(row.get("tmdb_id")) for row in movie_rows)
            if tmdb_id is not None
        }
        existing_movie_tmdb_ids = self._existing_movie_tmdb_ids(db, movie_tmdb_ids)
        known_source_movie_tmdb_ids = set(existing_movie_tmdb_ids)
        skipped_existing_movies = 0

        for row in movie_rows:
            tmdb_id = self._as_int(row.get("tmdb_id"))
            if tmdb_id is None:
                raise ExcelServiceError("Movie row is missing tmdb_id.")
            if tmdb_id in known_source_movie_tmdb_ids:
                skipped_existing_movies += 1
                continue

            movie = self._insert_movie_from_row(db, row, id_key="tmdb_id")
            source_movies[movie.tmdb_id] = movie
            imported_movie_ids.add(movie.id)
            known_source_movie_tmdb_ids.add(tmdb_id)

        similar_rows = self._sheet_rows(workbook, "similar_movies")
        similar_tmdb_ids = {
            similar_tmdb_id
            for similar_tmdb_id in (self._as_int(row.get("similar_tmdb_id")) for row in similar_rows)
            if similar_tmdb_id is not None
        }
        known_movie_tmdb_ids = set(known_source_movie_tmdb_ids)
        known_movie_tmdb_ids.update(source_movies.keys())
        existing_similar_tmdb_ids = self._existing_movie_tmdb_ids(db, similar_tmdb_ids)
        known_movie_tmdb_ids.update(existing_similar_tmdb_ids)

        for row in similar_rows:
            source_movie_tmdb_id = self._as_int(row.get("movie_tmdb_id"))
            similar_tmdb_id = self._as_int(row.get("similar_tmdb_id"))
            if source_movie_tmdb_id not in source_movies or similar_tmdb_id is None:
                continue
            if similar_tmdb_id in known_movie_tmdb_ids:
                continue

            similar_movie = self._insert_movie_from_row(db, row, id_key="similar_tmdb_id")
            imported_movie_ids.add(similar_movie.id)
            known_movie_tmdb_ids.add(similar_tmdb_id)

        genre_map: dict[int, list[Genre]] = {movie.id: [] for movie in source_movies.values()}
        for row in self._sheet_rows(workbook, "genres"):
            movie_tmdb_id = self._as_int(row.get("movie_tmdb_id"))
            genre_name = row.get("genre_name")
            if movie_tmdb_id not in source_movies or not genre_name:
                continue

            genre = None
            genre_tmdb_id = self._as_int(row.get("genre_tmdb_id"))
            if genre_tmdb_id is not None:
                genre = db.query(Genre).filter(Genre.tmdb_id == genre_tmdb_id).first()
            if genre is None:
                genre = db.query(Genre).filter(Genre.name == genre_name).first()
            if genre is None:
                genre = Genre(tmdb_id=genre_tmdb_id, name=genre_name)
                db.add(genre)
                db.flush()
            genre_map[source_movies[movie_tmdb_id].id].append(genre)

        for movie in source_movies.values():
            movie.genres = list(dict.fromkeys(genre_map[movie.id]))

        source_movie_ids = [movie.id for movie in source_movies.values()]
        for movie_id in source_movie_ids:
            db.execute(delete(movie_actors).where(movie_actors.c.movie_id == movie_id))

        actor_rows_to_insert: list[dict[str, Any]] = []
        for row in self._sheet_rows(workbook, "actors"):
            movie_tmdb_id = self._as_int(row.get("movie_tmdb_id"))
            actor_name = row.get("actor_name")
            if movie_tmdb_id not in source_movies or not actor_name:
                continue

            actor = None
            actor_tmdb_id = self._as_int(row.get("actor_tmdb_id"))
            if actor_tmdb_id is not None:
                actor = db.query(Actor).filter(Actor.tmdb_id == actor_tmdb_id).first()
            if actor is None:
                actor = db.query(Actor).filter(Actor.name == actor_name).first()
            if actor is None:
                actor = Actor(tmdb_id=actor_tmdb_id)
                db.add(actor)

            actor.name = actor_name
            actor.photo_url = row.get("photo_url")
            actor.popularity = self._as_float(row.get("popularity"))
            db.flush()

            actor_rows_to_insert.append(
                {
                    "movie_id": source_movies[movie_tmdb_id].id,
                    "actor_id": actor.id,
                    "character_name": row.get("character_name"),
                    "cast_order": self._as_int(row.get("cast_order")),
                },
            )

        if actor_rows_to_insert:
            db.execute(insert(movie_actors), actor_rows_to_insert)

        director_map: dict[int, list[Director]] = {movie.id: [] for movie in source_movies.values()}
        for row in self._sheet_rows(workbook, "directors"):
            movie_tmdb_id = self._as_int(row.get("movie_tmdb_id"))
            director_name = row.get("director_name")
            if movie_tmdb_id not in source_movies or not director_name:
                continue

            director = None
            director_tmdb_id = self._as_int(row.get("director_tmdb_id"))
            if director_tmdb_id is not None:
                director = db.query(Director).filter(Director.tmdb_id == director_tmdb_id).first()
            if director is None:
                director = db.query(Director).filter(Director.name == director_name).first()
            if director is None:
                director = Director(tmdb_id=director_tmdb_id)
                db.add(director)

            director.name = director_name
            director.photo_url = row.get("photo_url")
            db.flush()
            director_map[source_movies[movie_tmdb_id].id].append(director)

        for movie in source_movies.values():
            movie.directors = list(dict.fromkeys(director_map[movie.id]))

        keyword_map: dict[int, list[Keyword]] = {movie.id: [] for movie in source_movies.values()}
        for row in self._sheet_rows(workbook, "keywords"):
            movie_tmdb_id = self._as_int(row.get("movie_tmdb_id"))
            keyword_name = row.get("keyword_name")
            if movie_tmdb_id not in source_movies or not keyword_name:
                continue

            keyword = None
            keyword_tmdb_id = self._as_int(row.get("keyword_tmdb_id"))
            if keyword_tmdb_id is not None:
                keyword = db.query(Keyword).filter(Keyword.tmdb_id == keyword_tmdb_id).first()
            if keyword is None:
                keyword = db.query(Keyword).filter(Keyword.name == keyword_name).first()
            if keyword is None:
                keyword = Keyword(tmdb_id=keyword_tmdb_id, name=keyword_name)
                db.add(keyword)
                db.flush()
            keyword_map[source_movies[movie_tmdb_id].id].append(keyword)

        for movie in source_movies.values():
            movie.keywords = list(dict.fromkeys(keyword_map[movie.id]))

        similar_map: dict[int, list[Movie]] = defaultdict(list)
        similar_lookup = {
            movie.tmdb_id: movie
            for movie in db.query(Movie).filter(Movie.tmdb_id.in_(similar_tmdb_ids)).all()
        }
        for row in similar_rows:
            movie_tmdb_id = self._as_int(row.get("movie_tmdb_id"))
            similar_tmdb_id = self._as_int(row.get("similar_tmdb_id"))
            if movie_tmdb_id not in source_movies or similar_tmdb_id is None:
                continue

            similar_movie = similar_lookup.get(similar_tmdb_id)
            source_movie = source_movies[movie_tmdb_id]
            if similar_movie is None or similar_movie.id == source_movie.id:
                continue
            similar_map[source_movie.id].append(similar_movie)

        for movie in source_movies.values():
            movie.similar_movies = list(dict.fromkeys(similar_map[movie.id]))

        db.commit()
        return {
            "file_path": source_name,
            "processed_movies": len(movie_rows),
            "imported_movies": len(source_movies),
            "upserted_movies": len(imported_movie_ids),
            "skipped_existing_movies": skipped_existing_movies,
            "imported_movie_ids": list(imported_movie_ids),
            "imported_at": datetime.utcnow().isoformat(),
        }

    def import_tmdb_from_excel(self, db: Session, file_path: str) -> dict[str, Any]:
        _, load_workbook = self._load_openpyxl()
        resolved_path = self._resolve_import_path(file_path)
        workbook = load_workbook(resolved_path, data_only=True)
        return self._import_tmdb_from_workbook(db, workbook, str(resolved_path))

    def import_tmdb_from_bytes(
        self,
        db: Session,
        content: bytes,
        file_name: str,
    ) -> dict[str, Any]:
        _, load_workbook = self._load_openpyxl()
        workbook = load_workbook(BytesIO(content), data_only=True)
        return self._import_tmdb_from_workbook(db, workbook, file_name)


excel_service = ExcelService()
